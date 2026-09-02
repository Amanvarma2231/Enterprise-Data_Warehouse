import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
from rich.console import Console

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from src.config import DOCS_DIR, METADATA_DIR

console = Console(soft_wrap=True)

# Complete enterprise metadata definition catalog (60+ attributes)
METADATA_ENTRIES = [
    # dim_customer
    {"asset_name": "Customer Surrogate Key", "table_name": "dim_customer", "column_name": "customer_key", "data_type": "BIGINT", "business_definition": "System generated unique surrogate key for warehouse customer dimension", "source_system": "Internal DW", "source_column": "ROW_NUMBER()", "sensitivity": "INTERNAL", "owner": "Data Engineering", "refresh_frequency": "Daily Batch", "quality_rule": "PK / NOT NULL / UNIQUE"},
    {"asset_name": "Natural Customer ID", "table_name": "dim_customer", "column_name": "customer_id", "data_type": "BIGINT", "business_definition": "Primary customer identifier assigned by CRM / ERP transactional system", "source_system": "CRM / POS", "source_column": "customers.customer_id", "sensitivity": "INTERNAL", "owner": "Customer Ops", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / > 0"},
    {"asset_name": "First Name", "table_name": "dim_customer", "column_name": "first_name", "data_type": "VARCHAR(100)", "business_definition": "Given name of the registered customer", "source_system": "CRM", "source_column": "customers.first_name", "sensitivity": "CONFIDENTIAL (PII)", "owner": "Customer Ops", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / TRIM"},
    {"asset_name": "Last Name", "table_name": "dim_customer", "column_name": "last_name", "data_type": "VARCHAR(100)", "business_definition": "Family surname of the registered customer", "source_system": "CRM", "source_column": "customers.last_name", "sensitivity": "CONFIDENTIAL (PII)", "owner": "Customer Ops", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / TRIM"},
    {"asset_name": "Full Customer Name", "table_name": "dim_customer", "column_name": "full_name", "data_type": "VARCHAR(200)", "business_definition": "Concatenated full name for reporting and customer communication", "source_system": "Internal DW", "source_column": "first_name || ' ' || last_name", "sensitivity": "CONFIDENTIAL (PII)", "owner": "Customer Ops", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL"},
    {"asset_name": "Email Address", "table_name": "dim_customer", "column_name": "email", "data_type": "VARCHAR(255)", "business_definition": "Primary contact electronic mail address", "source_system": "CRM", "source_column": "customers.email", "sensitivity": "CONFIDENTIAL (PII)", "owner": "Marketing", "refresh_frequency": "Daily Batch", "quality_rule": "Regex Email Pattern: %@%.%"},
    {"asset_name": "Phone Number", "table_name": "dim_customer", "column_name": "phone", "data_type": "VARCHAR(50)", "business_definition": "Primary contact telephone / mobile number", "source_system": "CRM", "source_column": "customers.phone", "sensitivity": "CONFIDENTIAL (PII)", "owner": "Customer Ops", "refresh_frequency": "Daily Batch", "quality_rule": "Length <= 15"},
    {"asset_name": "City", "table_name": "dim_customer", "column_name": "city", "data_type": "VARCHAR(100)", "business_definition": "Municipal billing / residential city", "source_system": "CRM", "source_column": "customers.city", "sensitivity": "INTERNAL", "owner": "Sales Analytics", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL"},
    {"asset_name": "State / Province", "table_name": "dim_customer", "column_name": "state", "data_type": "VARCHAR(100)", "business_definition": "Administrative provincial territory", "source_system": "CRM", "source_column": "customers.state", "sensitivity": "INTERNAL", "owner": "Sales Analytics", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL"},
    {"asset_name": "Country", "table_name": "dim_customer", "column_name": "country", "data_type": "VARCHAR(100)", "business_definition": "Sovereign country name of registration", "source_system": "CRM", "source_column": "customers.country", "sensitivity": "PUBLIC", "owner": "Sales Analytics", "refresh_frequency": "Daily Batch", "quality_rule": "DEFAULT 'India'"},
    {"asset_name": "Postal Code", "table_name": "dim_customer", "column_name": "postal_code", "data_type": "VARCHAR(30)", "business_definition": "Postal / ZIP geographical routing code", "source_system": "CRM", "source_column": "customers.postal_code", "sensitivity": "INTERNAL", "owner": "Logistics", "refresh_frequency": "Daily Batch", "quality_rule": "Standard Format"},
    {"asset_name": "Customer Segment", "table_name": "dim_customer", "column_name": "segment", "data_type": "VARCHAR(50)", "business_definition": "Loyalty tier classification: Regular, Premium, VIP, Corporate", "source_system": "CRM / Marketing", "source_column": "customers.segment", "sensitivity": "INTERNAL", "owner": "Marketing", "refresh_frequency": "Daily Batch", "quality_rule": "IN ('Regular','Premium','VIP','Corporate','Occasional')"},
    {"asset_name": "Registration Date", "table_name": "dim_customer", "column_name": "registration_date", "data_type": "DATE", "business_definition": "Date when customer account was created", "source_system": "CRM", "source_column": "customers.registration_date", "sensitivity": "INTERNAL", "owner": "Customer Ops", "refresh_frequency": "Daily Batch", "quality_rule": "<= CURRENT_DATE"},
    {"asset_name": "Active Indicator", "table_name": "dim_customer", "column_name": "is_active", "data_type": "BOOLEAN", "business_definition": "Active profile flag (True/False)", "source_system": "CRM", "source_column": "customers.is_active", "sensitivity": "INTERNAL", "owner": "Customer Ops", "refresh_frequency": "Daily Batch", "quality_rule": "BOOLEAN"},

    # dim_product
    {"asset_name": "Product Surrogate Key", "table_name": "dim_product", "column_name": "product_key", "data_type": "BIGINT", "business_definition": "System generated unique surrogate key for product dimension", "source_system": "Internal DW", "source_column": "ROW_NUMBER()", "sensitivity": "INTERNAL", "owner": "Data Engineering", "refresh_frequency": "Daily Batch", "quality_rule": "PK / NOT NULL / UNIQUE"},
    {"asset_name": "Natural Product ID", "table_name": "dim_product", "column_name": "product_id", "data_type": "BIGINT", "business_definition": "Source system catalog product ID", "source_system": "Catalog ERP", "source_column": "products.product_id", "sensitivity": "INTERNAL", "owner": "Merchandising", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / > 0"},
    {"asset_name": "Stock Keeping Unit", "table_name": "dim_product", "column_name": "sku", "data_type": "VARCHAR(100)", "business_definition": "Global unique merchandise identification code", "source_system": "Catalog ERP", "source_column": "products.sku", "sensitivity": "INTERNAL", "owner": "Inventory Ops", "refresh_frequency": "Daily Batch", "quality_rule": "UNIQUE / NOT NULL"},
    {"asset_name": "Product Name", "table_name": "dim_product", "column_name": "product_name", "data_type": "VARCHAR(255)", "business_definition": "Official catalog merchandising title of the product", "source_system": "Catalog ERP", "source_column": "products.product_name", "sensitivity": "PUBLIC", "owner": "Merchandising", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / LENGTH > 2"},
    {"asset_name": "Product Category", "table_name": "dim_product", "column_name": "category", "data_type": "VARCHAR(100)", "business_definition": "Top-level department classification (e.g., Electronics, Apparel)", "source_system": "Catalog ERP", "source_column": "products.category", "sensitivity": "PUBLIC", "owner": "Merchandising", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL"},
    {"asset_name": "Product Subcategory", "table_name": "dim_product", "column_name": "subcategory", "data_type": "VARCHAR(100)", "business_definition": "Granular department subcategory (e.g., Laptops, Footwear)", "source_system": "Catalog ERP", "source_column": "products.subcategory", "sensitivity": "PUBLIC", "owner": "Merchandising", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL"},
    {"asset_name": "Unit Cost", "table_name": "dim_product", "column_name": "unit_cost", "data_type": "DECIMAL(12,2)", "business_definition": "Procurement / manufacturing acquisition cost per unit (INR)", "source_system": "Procurement ERP", "source_column": "products.unit_cost", "sensitivity": "RESTRICTED", "owner": "Finance", "refresh_frequency": "Daily Batch", "quality_rule": "> 0.00"},
    {"asset_name": "Unit Price", "table_name": "dim_product", "column_name": "unit_price", "data_type": "DECIMAL(12,2)", "business_definition": "Standard suggested retail selling price per unit (INR)", "source_system": "Catalog ERP", "source_column": "products.unit_price", "sensitivity": "PUBLIC", "owner": "Pricing Strategy", "refresh_frequency": "Daily Batch", "quality_rule": "unit_price >= unit_cost"},
    {"asset_name": "Profit Margin %", "table_name": "dim_product", "column_name": "profit_margin_pct", "data_type": "DECIMAL(6,2)", "business_definition": "Base profit margin percentage ((price - cost) / price) * 100", "source_system": "Internal DW", "source_column": "((unit_price - unit_cost)/unit_price)*100", "sensitivity": "CONFIDENTIAL", "owner": "Finance", "refresh_frequency": "Daily Batch", "quality_rule": "BETWEEN 0 AND 100"},
    {"asset_name": "Price Tier", "table_name": "dim_product", "column_name": "price_tier", "data_type": "VARCHAR(50)", "business_definition": "Pricing bracket: Budget, Mid-Range, Premium, Luxury", "source_system": "Internal DW", "source_column": "CASE unit_price", "sensitivity": "INTERNAL", "owner": "Pricing Strategy", "refresh_frequency": "Daily Batch", "quality_rule": "IN ('Budget','Mid-Range','Premium','Luxury')"},
    {"asset_name": "Reorder Level", "table_name": "dim_product", "column_name": "reorder_level", "data_type": "INTEGER", "business_definition": "Inventory minimum threshold before triggering reorder", "source_system": "Inventory ERP", "source_column": "products.reorder_level", "sensitivity": "INTERNAL", "owner": "Supply Chain", "refresh_frequency": "Daily Batch", "quality_rule": ">= 0"},
    {"asset_name": "Discontinued Flag", "table_name": "dim_product", "column_name": "is_discontinued", "data_type": "BOOLEAN", "business_definition": "Lifecycle status indicating if item has been discontinued", "source_system": "Catalog ERP", "source_column": "products.is_discontinued", "sensitivity": "INTERNAL", "owner": "Merchandising", "refresh_frequency": "Daily Batch", "quality_rule": "BOOLEAN"},

    # dim_store
    {"asset_name": "Store Surrogate Key", "table_name": "dim_store", "column_name": "store_key", "data_type": "BIGINT", "business_definition": "System generated surrogate key for retail store dimension", "source_system": "Internal DW", "source_column": "ROW_NUMBER()", "sensitivity": "INTERNAL", "owner": "Data Engineering", "refresh_frequency": "Daily Batch", "quality_rule": "PK / NOT NULL / UNIQUE"},
    {"asset_name": "Natural Store ID", "table_name": "dim_store", "column_name": "store_id", "data_type": "BIGINT", "business_definition": "Source store branch unique identifier", "source_system": "Retail POS ERP", "source_column": "stores.store_id", "sensitivity": "INTERNAL", "owner": "Retail Ops", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / > 0"},
    {"asset_name": "Store Trading Name", "table_name": "dim_store", "column_name": "store_name", "data_type": "VARCHAR(255)", "business_definition": "Official retail outlet location display name", "source_system": "Retail POS ERP", "source_column": "stores.store_name", "sensitivity": "PUBLIC", "owner": "Retail Ops", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL"},
    {"asset_name": "Store Format Type", "table_name": "dim_store", "column_name": "store_type", "data_type": "VARCHAR(100)", "business_definition": "Retail footprint format: Flagship, Standard, Outlet, Online", "source_system": "Retail POS ERP", "source_column": "stores.store_type", "sensitivity": "INTERNAL", "owner": "Retail Ops", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL"},
    {"asset_name": "Channel Grouping", "table_name": "dim_store", "column_name": "channel_group", "data_type": "VARCHAR(50)", "business_definition": "Omnichannel division: Physical Retail vs Digital Channel", "source_system": "Internal DW", "source_column": "CASE store_type", "sensitivity": "INTERNAL", "owner": "Executive Strategy", "refresh_frequency": "Daily Batch", "quality_rule": "IN ('Physical Retail', 'Digital Channel')"},
    {"asset_name": "Sales Region", "table_name": "dim_store", "column_name": "region", "data_type": "VARCHAR(100)", "business_definition": "Geographic retail operating territory (North, South, East, West, Central)", "source_system": "Retail POS ERP", "source_column": "stores.region", "sensitivity": "INTERNAL", "owner": "Sales Analytics", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL"},
    {"asset_name": "Store Square Footage", "table_name": "dim_store", "column_name": "square_feet", "data_type": "INTEGER", "business_definition": "Total commercial floor area in square feet", "source_system": "Facilities ERP", "source_column": "stores.square_feet", "sensitivity": "INTERNAL", "owner": "Facilities", "refresh_frequency": "Daily Batch", "quality_rule": ">= 0"},
    {"asset_name": "Store Opening Date", "table_name": "dim_store", "column_name": "opened_date", "data_type": "DATE", "business_definition": "Grand opening commercial launch date", "source_system": "Retail POS ERP", "source_column": "stores.opened_date", "sensitivity": "INTERNAL", "owner": "Retail Ops", "refresh_frequency": "Daily Batch", "quality_rule": "<= CURRENT_DATE"},

    # dim_date
    {"asset_name": "Date Surrogate Key", "table_name": "dim_date", "column_name": "date_key", "data_type": "INTEGER", "business_definition": "Smart integer calendar key in YYYYMMDD format", "source_system": "Conformed Calendar", "source_column": "strftime('%Y%m%d')", "sensitivity": "PUBLIC", "owner": "Data Engineering", "refresh_frequency": "Static Pre-populated", "quality_rule": "PK / NOT NULL / 8 digits"},
    {"asset_name": "Full Date", "table_name": "dim_date", "column_name": "full_date", "data_type": "DATE", "business_definition": "Standard calendar ISO date", "source_system": "Conformed Calendar", "source_column": "calendar_date", "sensitivity": "PUBLIC", "owner": "Data Engineering", "refresh_frequency": "Static Pre-populated", "quality_rule": "UNIQUE / NOT NULL"},
    {"asset_name": "Year Month", "table_name": "dim_date", "column_name": "year_month", "data_type": "VARCHAR(7)", "business_definition": "Period identifier in YYYY-MM format", "source_system": "Conformed Calendar", "source_column": "strftime('%Y-%m')", "sensitivity": "PUBLIC", "owner": "Data Engineering", "refresh_frequency": "Static Pre-populated", "quality_rule": "Regex ^[0-9]{4}-[0-9]{2}$"},
    {"asset_name": "Fiscal Quarter", "table_name": "dim_date", "column_name": "fiscal_quarter", "data_type": "VARCHAR(10)", "business_definition": "Corporate financial reporting quarter (e.g., FY26-Q2)", "source_system": "Conformed Calendar", "source_column": "Fiscal calculation", "sensitivity": "PUBLIC", "owner": "Finance", "refresh_frequency": "Static Pre-populated", "quality_rule": "NOT NULL"},

    # fact_sales
    {"asset_name": "Sales Fact Surrogate Key", "table_name": "fact_sales", "column_name": "sales_key", "data_type": "BIGINT", "business_definition": "Surrogate primary key for individual order line item in fact table", "source_system": "Internal DW", "source_column": "ROW_NUMBER()", "sensitivity": "INTERNAL", "owner": "Data Engineering", "refresh_frequency": "Daily Batch", "quality_rule": "PK / NOT NULL / UNIQUE"},
    {"asset_name": "Degenerate Order ID", "table_name": "fact_sales", "column_name": "order_id", "data_type": "BIGINT", "business_definition": "Transaction order identifier from POS (Degenerate Dimension)", "source_system": "POS / E-Commerce", "source_column": "orders.order_id", "sensitivity": "INTERNAL", "owner": "Sales Analytics", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / > 0"},
    {"asset_name": "Degenerate Order Item ID", "table_name": "fact_sales", "column_name": "order_item_id", "data_type": "BIGINT", "business_definition": "Line item identifier from order manifest", "source_system": "POS / E-Commerce", "source_column": "order_items.order_item_id", "sensitivity": "INTERNAL", "owner": "Sales Analytics", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / > 0"},
    {"asset_name": "Customer Foreign Key", "table_name": "fact_sales", "column_name": "customer_key", "data_type": "BIGINT", "business_definition": "Referential surrogate key connecting to dim_customer", "source_system": "Internal DW", "source_column": "dim_customer.customer_key", "sensitivity": "INTERNAL", "owner": "Data Engineering", "refresh_frequency": "Daily Batch", "quality_rule": "FK -> dim_customer.customer_key"},
    {"asset_name": "Product Foreign Key", "table_name": "fact_sales", "column_name": "product_key", "data_type": "BIGINT", "business_definition": "Referential surrogate key connecting to dim_product", "source_system": "Internal DW", "source_column": "dim_product.product_key", "sensitivity": "INTERNAL", "owner": "Data Engineering", "refresh_frequency": "Daily Batch", "quality_rule": "FK -> dim_product.product_key"},
    {"asset_name": "Store Foreign Key", "table_name": "fact_sales", "column_name": "store_key", "data_type": "BIGINT", "business_definition": "Referential surrogate key connecting to dim_store", "source_system": "Internal DW", "source_column": "dim_store.store_key", "sensitivity": "INTERNAL", "owner": "Data Engineering", "refresh_frequency": "Daily Batch", "quality_rule": "FK -> dim_store.store_key"},
    {"asset_name": "Date Foreign Key", "table_name": "fact_sales", "column_name": "date_key", "data_type": "INTEGER", "business_definition": "Referential smart key connecting to dim_date", "source_system": "Internal DW", "source_column": "dim_date.date_key", "sensitivity": "INTERNAL", "owner": "Data Engineering", "refresh_frequency": "Daily Batch", "quality_rule": "FK -> dim_date.date_key"},
    {"asset_name": "Quantity Sold", "table_name": "fact_sales", "column_name": "quantity", "data_type": "INTEGER", "business_definition": "Units of product sold in transaction line", "source_system": "POS / E-Commerce", "source_column": "order_items.quantity", "sensitivity": "INTERNAL", "owner": "Sales Analytics", "refresh_frequency": "Daily Batch", "quality_rule": "quantity > 0"},
    {"asset_name": "Selling Unit Price", "table_name": "fact_sales", "column_name": "unit_price", "data_type": "DECIMAL(12,2)", "business_definition": "Actual selling price per unit at time of sale", "source_system": "POS / E-Commerce", "source_column": "order_items.unit_price", "sensitivity": "INTERNAL", "owner": "Pricing Strategy", "refresh_frequency": "Daily Batch", "quality_rule": "unit_price > 0.00"},
    {"asset_name": "Gross Sales Amount", "table_name": "fact_sales", "column_name": "gross_sales_amount", "data_type": "DECIMAL(14,2)", "business_definition": "Gross monetary line total = quantity * unit_price", "source_system": "Internal DW", "source_column": "quantity * unit_price", "sensitivity": "INTERNAL", "owner": "Finance", "refresh_frequency": "Daily Batch", "quality_rule": "gross_sales_amount > 0.00"},
    {"asset_name": "Promotional Discount", "table_name": "fact_sales", "column_name": "discount_amount", "data_type": "DECIMAL(14,2)", "business_definition": "Total discount deduction applied to line item", "source_system": "POS / Promotions", "source_column": "order_items.discount", "sensitivity": "INTERNAL", "owner": "Finance", "refresh_frequency": "Daily Batch", "quality_rule": "discount_amount >= 0.00"},
    {"asset_name": "Net Sales Amount", "table_name": "fact_sales", "column_name": "net_sales_amount", "data_type": "DECIMAL(14,2)", "business_definition": "Net revenue earned after promotional discount = gross - discount", "source_system": "Internal DW", "source_column": "gross_sales_amount - discount_amount", "sensitivity": "INTERNAL", "owner": "Finance", "refresh_frequency": "Daily Batch", "quality_rule": "net_sales_amount >= 0.00"},
    {"asset_name": "Cost of Goods Sold (COGS)", "table_name": "fact_sales", "column_name": "cost_amount", "data_type": "DECIMAL(14,2)", "business_definition": "Total procurement cost = quantity * unit_cost", "source_system": "Internal DW", "source_column": "quantity * dim_product.unit_cost", "sensitivity": "RESTRICTED", "owner": "Finance", "refresh_frequency": "Daily Batch", "quality_rule": "cost_amount >= 0.00"},
    {"asset_name": "Gross Profit Amount", "table_name": "fact_sales", "column_name": "gross_profit_amount", "data_type": "DECIMAL(14,2)", "business_definition": "Gross margin contribution = net_sales_amount - cost_amount", "source_system": "Internal DW", "source_column": "net_sales_amount - cost_amount", "sensitivity": "CONFIDENTIAL", "owner": "Finance", "refresh_frequency": "Daily Batch", "quality_rule": "Numeric"},
    {"asset_name": "Profit Margin %", "table_name": "fact_sales", "column_name": "profit_margin_pct", "data_type": "DECIMAL(6,2)", "business_definition": "Realized gross profit margin percentage on transaction line", "source_system": "Internal DW", "source_column": "(gross_profit / net_sales) * 100", "sensitivity": "CONFIDENTIAL", "owner": "Finance", "refresh_frequency": "Daily Batch", "quality_rule": "Numeric Percentage"},

    # fact_payments
    {"asset_name": "Payment Fact Key", "table_name": "fact_payments", "column_name": "payment_key", "data_type": "BIGINT", "business_definition": "Surrogate primary key for payment transaction in warehouse", "source_system": "Internal DW", "source_column": "ROW_NUMBER()", "sensitivity": "INTERNAL", "owner": "Data Engineering", "refresh_frequency": "Daily Batch", "quality_rule": "PK / NOT NULL / UNIQUE"},
    {"asset_name": "Natural Payment ID", "table_name": "fact_payments", "column_name": "payment_id", "data_type": "BIGINT", "business_definition": "Payment gateway transactional settlement ID", "source_system": "Payment Gateway", "source_column": "payments.payment_id", "sensitivity": "INTERNAL", "owner": "Treasury Ops", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / > 0"},
    {"asset_name": "Payment Tender Method", "table_name": "fact_payments", "column_name": "payment_method", "data_type": "VARCHAR(50)", "business_definition": "Payment instrument: UPI, Credit Card, Debit Card, NetBanking, Cash", "source_system": "Payment Gateway", "source_column": "payments.payment_method", "sensitivity": "INTERNAL", "owner": "Treasury Ops", "refresh_frequency": "Daily Batch", "quality_rule": "IN ('UPI','Credit Card','Debit Card','NetBanking','PayPal','Cash on Delivery')"},
    {"asset_name": "Payment Settlement Status", "table_name": "fact_payments", "column_name": "payment_status", "data_type": "VARCHAR(50)", "business_definition": "Settlement outcome: Success, Failed, Refunded, Pending", "source_system": "Payment Gateway", "source_column": "payments.payment_status", "sensitivity": "INTERNAL", "owner": "Treasury Ops", "refresh_frequency": "Daily Batch", "quality_rule": "IN ('Success','Failed','Refunded','Pending')"},
    {"asset_name": "Settled Payment Amount", "table_name": "fact_payments", "column_name": "payment_amount", "data_type": "DECIMAL(14,2)", "business_definition": "Actual transacted currency amount settled through gateway", "source_system": "Payment Gateway", "source_column": "payments.payment_amount", "sensitivity": "INTERNAL", "owner": "Finance", "refresh_frequency": "Daily Batch", "quality_rule": "payment_amount > 0.00"},
    {"asset_name": "Bank Transaction Ref", "table_name": "fact_payments", "column_name": "transaction_ref", "data_type": "VARCHAR(100)", "business_definition": "External banking authorization reference code", "source_system": "Payment Gateway", "source_column": "payments.transaction_ref", "sensitivity": "CONFIDENTIAL", "owner": "Treasury Ops", "refresh_frequency": "Daily Batch", "quality_rule": "NOT NULL / Standard Prefix"},
]


def generate_metadata_repository() -> pd.DataFrame:
    """Generate and save formal metadata repository CSV."""
    METADATA_DIR.mkdir(parents=True, exist_ok=True)
    df_meta = pd.DataFrame(METADATA_ENTRIES)
    csv_path = METADATA_DIR / "metadata.csv"
    df_meta.to_csv(csv_path, index=False)
    console.print(f"[bold green]✔ Saved metadata repository -> {csv_path.name} ({len(df_meta)} attributes)[/bold green]")
    return df_meta


def generate_data_dictionary_artifacts(df_meta: pd.DataFrame):
    """Generate Data Dictionary in Markdown and multi-tab formatted Excel format."""
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    
    # 1. Markdown Artifact
    md_path = DOCS_DIR / "data_dictionary.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# RetailSphere Enterprise Data Warehouse - Data Dictionary\n\n")
        f.write("This document provides business and technical definitions, source lineage, data types, ")
        f.write("security classifications, and quality validation rules for all warehouse entities.\n\n")
        
        tables = df_meta["table_name"].unique()
        for t in tables:
            f.write(f"## Table: `{t}`\n\n")
            df_t = df_meta[df_meta["table_name"] == t][[
                "column_name", "data_type", "business_definition", "sensitivity", "source_system", "quality_rule"
            ]]
            f.write(df_t.to_markdown(index=False))
            f.write("\n\n---\n\n")
            
    console.print(f"[bold green]✔ Saved Markdown Data Dictionary -> {md_path.name}[/bold green]")

    # 2. Excel Artifact with Professional Styling
    xlsx_path = DOCS_DIR / "data_dictionary.xlsx"
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Overview & Governance"
    ws_summary.sheet_properties.tabColor = "104E8B"
    
    ws_summary["A1"] = "RetailSphere Enterprise Data Warehouse - Data Dictionary"
    ws_summary["A1"].font = Font(name="Calibri", size=16, bold=True, color="104E8B")
    ws_summary["A2"] = f"Total Documented Data Assets: {len(df_meta)} columns across 5 core dimensional models"
    ws_summary["A2"].font = Font(name="Calibri", size=11, italic=True)
    
    headers = ["Asset Name", "Table Name", "Column Name", "Data Type", "Business Definition", "Source System", "Source Column", "Sensitivity", "Owner", "Refresh Frequency", "Quality Rule"]
    
    # All Data Assets Sheet
    ws_catalog = wb.create_sheet(title="Full Data Catalog")
    ws_catalog.sheet_properties.tabColor = "008B8B"
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    ws_catalog.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws_catalog.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx, entry in enumerate(METADATA_ENTRIES, start=2):
        ws_catalog.append([
            entry["asset_name"], entry["table_name"], entry["column_name"], entry["data_type"],
            entry["business_definition"], entry["source_system"], entry["source_column"],
            entry["sensitivity"], entry["owner"], entry["refresh_frequency"], entry["quality_rule"]
        ])
        for col_idx in range(1, len(headers) + 1):
            cell = ws_catalog.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if entry["sensitivity"] == "CONFIDENTIAL (PII)" and col_idx == 8:
                cell.fill = PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid")
                cell.font = Font(name="Calibri", size=10, bold=True, color="900000")

    # Auto-adjust column widths
    for sheet in [ws_catalog]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb.save(xlsx_path)
    console.print(f"[bold green]✔ Saved Formatted Excel Data Dictionary -> {xlsx_path.name}[/bold green]")


def generate_governance_and_lineage_docs():
    """Generate markdown files for Governance, Standards, Lineage, and Semantic Layer."""
    
    # 1. Data Lineage Document
    lineage_path = DOCS_DIR / "data-lineage.md"
    with open(lineage_path, "w", encoding="utf-8") as f:
        f.write("""# RetailSphere Data Platform - End-to-End Data Lineage

## 1. High-Level Pipeline Lineage Architecture

```mermaid
flowchart TD
    subgraph S1 [Operational Data Sources]
        C[customers.csv]
        P[products.csv]
        S[stores.csv]
        O[orders.csv]
        OI[order_items.csv]
        PY[payments.csv]
    end

    subgraph S2 [Ingestion & Staging Layer]
        SC[staging.stg_customers]
        SP[staging.stg_products]
        SS[staging.stg_stores]
        SO[staging.stg_orders]
        SOI[staging.stg_order_items]
        SPY[staging.stg_payments]
    end

    subgraph S3 [Data Quality & Quarantine Layer]
        DQ{Data Quality Engine\\n10 Rule Checks}
        Q1[quarantine_orders]
        Q2[quarantine_order_items]
        Q3[quarantine_customers]
    end

    subgraph S4 [Dimensional Star Schema Warehouse]
        DC[dim_customer]
        DP[dim_product]
        DS[dim_store]
        DD[dim_date]
        FS[fact_sales]
        FP[fact_payments]
    end

    subgraph S5 [Analytical Marts & Semantic Layer]
        M1[mart_monthly_store_performance]
        M2[mart_customer_rfm]
        BI[Executive BI Dashboard]
        BQ[Google BigQuery Cloud DW]
    end

    C --> SC
    P --> SP
    S --> SS
    O --> SO
    OI --> SOI
    PY --> SPY

    SC & SP & SS & SO & SOI & SPY --> DQ
    DQ -- "Invalid Records" --> Q1 & Q2 & Q3
    DQ -- "Valid Cleansed Records" --> DC & DP & DS & DD

    SC --> DC
    SP --> DP
    SS --> DS
    SO & SOI & DC & DP & DS & DD --> FS
    SPY & SO & DC & DD --> FP

    FS & DS & DD --> M1
    FS & DC --> M2
    M1 & M2 & FS --> BI
    FS & DC & DP & DS & DD --> BQ
```

## 2. Table-to-Table Dependency Matrix

| Target Table | Source Entity | Transformation Type | Business Logic & Surrogate Key Generation |
| :--- | :--- | :--- | :--- |
| `dim_customer` | `staging.stg_customers` | Cleansing, Deduplication, SCD Type 1/2 | Deduped on `customer_id` keeping newest registration; surrogate key `customer_key` generated via window ranking |
| `dim_product` | `staging.stg_products` | Enrichment, Pricing Tiering | Added `profit_margin_pct` and classified into `Budget`, `Mid-Range`, `Premium`, `Luxury` tiers |
| `dim_store` | `staging.stg_stores` | Channel Grouping, Age Calculation | Segmented into `Physical Retail` vs `Digital Channel`; computed `store_age_years` |
| `dim_date` | Date Generator Algorithm | Conformed Calendar Dimension | Generates full calendar, fiscal periods (Indian FY April-March), weekend and holiday flags |
| `fact_sales` | `stg_orders`, `stg_order_items`, Dimensions | Fact Grain Joins & Financial Metrics | Grain: 1 row per order item. Evaluates `net_sales_amount`, `cost_amount`, `gross_profit_amount`, `profit_margin_pct` |
| `fact_payments` | `stg_payments`, `stg_orders`, `dim_customer` | Financial Reconciliation | Maps payment transactions, validates against order values, evaluates success rates |
| `mart_monthly_store_performance` | `fact_sales`, `dim_date`, `dim_store` | Periodic Monthly Snapshot Aggregation | Computes MoM revenue, order counts, gross profit, margin percentage per store and territory |
| `mart_customer_rfm` | `fact_sales`, `dim_customer` | Behavioral Customer Segmentation | Calculates Recency (days), Frequency (order count), Monetary (total spend), and assigns RFM tiers |
""")
    console.print(f"[bold green]✔ Saved Data Lineage -> {lineage_path.name}[/bold green]")

    # 2. Data Governance Document
    gov_path = DOCS_DIR / "governance.md"
    with open(gov_path, "w", encoding="utf-8") as f:
        f.write("""# RetailSphere Enterprise Data Governance Framework

## 1. Information Security & Data Classification Policy

RetailSphere enforces a 4-tier data classification hierarchy aligned with GDPR, CCPA, and ISO/IEC 27001 standards:

| Classification Tier | Definition & Scope | Examples | Access Controls & Encryption |
| :--- | :--- | :--- | :--- |
| **PUBLIC** | Information that can be freely shared externally without risk to the enterprise. | Product Catalog Names, Store Locations, Public Pricing | Unrestricted Read Access; SSL/TLS in transit |
| **INTERNAL** | Standard operational data intended for internal staff and analytics workloads. | Aggregated Sales, Store IDs, SKU metadata, General Metrics | Role-Based Access Control (RBAC); corporate network / VPN |
| **CONFIDENTIAL (PII)** | Personally Identifiable Information (PII) that directly identifies individual humans. | Customer First/Last Names, Email Addresses, Phone Numbers | Column-level encryption, Dynamic Data Masking, strict RBAC |
| **RESTRICTED** | Highly sensitive proprietary commercial metrics and secret credentials. | Unit Procurement Cost, Bank Authorization Hashes, Margin Secrets | Column-level hashing, multi-factor authorization, immutable audit logs |

---

## 2. PII Data Masking Standards (SQL Implementation)

For non-privileged analytical users, PII fields must be dynamically masked:

```sql
-- Dynamic Email Masking
CREATE VIEW analytics.v_dim_customer AS
SELECT 
    customer_key,
    customer_id,
    SUBSTRING(first_name, 1, 1) || '****' AS first_name,
    SUBSTRING(last_name, 1, 1) || '****' AS last_name,
    SUBSTRING(email, 1, 2) || '****@' || SPLIT_PART(email, '@', 2) AS masked_email,
    '***-***-' || RIGHT(phone, 4) AS masked_phone,
    city,
    state,
    country,
    segment
FROM warehouse.dim_customer;
```

---

## 3. Data Quality SLA & Quarantine Rules

- **Zero-Tolerance Rules (Quarantine Trigger):** Any record with a `NULL` Primary Key, orphaned Foreign Key, or negative quantity is immediately redirected to `data/quarantine/`.
- **Warning Rules (Audit Flag):** Missing optional attributes (postal code, store manager) trigger metric logs but do not block pipeline progression.
""")
    console.print(f"[bold green]✔ Saved Governance Framework -> {gov_path.name}[/bold green]")

    # 3. Modelling Standards Document
    standards_path = DOCS_DIR / "modelling-standards.md"
    with open(standards_path, "w", encoding="utf-8") as f:
        f.write("""# RetailSphere Enterprise Data Modelling Standards & Conventions

## 1. Naming Conventions

- **Schema Names:** Lowercase snake_case (`staging`, `warehouse`, `quarantine`, `analytics`).
- **Dimension Tables:** Prefix `dim_` (e.g., `dim_customer`, `dim_product`, `dim_store`, `dim_date`).
- **Fact Tables:** Prefix `fact_` (e.g., `fact_sales`, `fact_payments`).
- **Analytical Marts:** Prefix `mart_` (e.g., `mart_monthly_store_performance`, `mart_customer_rfm`).
- **Surrogate Keys:** Named `<entity>_key` with `BIGINT` or `INTEGER` data type.
- **Natural / Business Keys:** Named `<entity>_id` matching source system keys.
- **Audit Columns:** Preceded with underscore (e.g., `_loaded_at`, `_source_file`).

---

## 2. Dimensional Modelling Rules (Kimball Methodology)

1. **Grain Declaration:** Every Fact table must have an explicitly documented atomic grain.
   - `fact_sales`: One row per line item in a completed retail sales order.
   - `fact_payments`: One row per financial settlement attempt.
2. **Conformed Dimensions:** Dimensions like `dim_date`, `dim_customer`, and `dim_store` must be shared across all business processes to ensure drill-across query consistency.
3. **Surrogate Keys:** Natural keys from upstream systems must never be used as primary keys in the dimensional model. Surrogate keys shield the warehouse from upstream key reuse and support SCD Type 2 tracking.
4. **Slowly Changing Dimensions (SCD):**
   - Customer dimension implements **SCD Type 1** for rapid operational attribute updates with `row_effective_date`, `row_expiration_date`, and `is_current` flags to support seamless evolution to **SCD Type 2**.
""")
    console.print(f"[bold green]✔ Saved Modelling Standards -> {standards_path.name}[/bold green]")

    # 4. Semantic Layer Document
    semantic_path = DOCS_DIR / "semantic_layer.md"
    with open(semantic_path, "w", encoding="utf-8") as f:
        f.write("""# RetailSphere Semantic Layer & Metric Definitions

The Semantic Layer provides centralized, governed metric logic across BI tools, SQL queries, and Python notebooks to eliminate metric divergence.

## Core Governed Metrics

| Metric Name | Business Definition | Mathematical Formula | SQL Representation |
| :--- | :--- | :--- | :--- |
| **Gross Revenue** | Total merchandise sales value before any promotional discounts | `Σ (quantity * unit_price)` | `SUM(gross_sales_amount)` |
| **Promotional Deductions** | Total customer savings and promotional deductions | `Σ (discount_amount)` | `SUM(discount_amount)` |
| **Net Revenue** | Realized sales revenue after deducting discounts | `Gross Revenue - Promotional Deductions` | `SUM(net_sales_amount)` |
| **Cost of Goods Sold (COGS)** | Total inventory procurement cost for items sold | `Σ (quantity * unit_cost)` | `SUM(cost_amount)` |
| **Gross Profit** | Gross monetary margin contribution | `Net Revenue - COGS` | `SUM(gross_profit_amount)` |
| **Gross Profit Margin %** | Percentage of net revenue retained as gross profit | `(Gross Profit / Net Revenue) * 100` | `ROUND((SUM(gross_profit_amount) / NULLIF(SUM(net_sales_amount), 0)) * 100.0, 2)` |
| **Total Order Count** | Total distinct orders placed | `Count(Distinct order_id)` | `COUNT(DISTINCT order_id)` |
| **Average Order Value (AOV)** | Average revenue generated per distinct order | `Net Revenue / Total Order Count` | `ROUND(SUM(net_sales_amount) / NULLIF(COUNT(DISTINCT order_id), 0), 2)` |
| **Units per Transaction (UPT)** | Average quantity of items purchased per order | `Total Units Sold / Total Order Count` | `ROUND(SUM(quantity) * 1.0 / NULLIF(COUNT(DISTINCT order_id), 0), 2)` |
| **Payment Success Rate %** | Percentage of attempted payments that succeeded | `(Successful Payments / Total Payments) * 100` | `ROUND((COUNT(CASE WHEN is_successful THEN 1 END) * 100.0) / COUNT(*), 2)` |
""")
    console.print(f"[bold green]✔ Saved Semantic Layer -> {semantic_path.name}[/bold green]")


def run_governance_generation():
    """Execute complete governance and metadata build."""
    console.print("[bold cyan][*] Generating Enterprise Governance & Metadata Repository...[/bold cyan]")
    df_meta = generate_metadata_repository()
    generate_data_dictionary_artifacts(df_meta)
    generate_governance_and_lineage_docs()
    console.print("[bold green]✔ Governance & Metadata build complete![/bold green]")


if __name__ == "__main__":
    run_governance_generation()
